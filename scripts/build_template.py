# -*- coding: utf-8 -*-
"""Build scheduling-template.xlsx with sheets HƯỚNG DẪN, INPUT, TÍNH TOÁN, OUTPUT.

4-sheet model:
- HƯỚNG DẪN   — instructions
- INPUT       — all data entry (INPUT 1-5)
- TÍNH TOÁN   — compute block with cross-sheet formulas referencing INPUT!
- OUTPUT      — final schedule with cross-sheet formulas referencing TÍNH TOÁN!
"""

from __future__ import annotations

import os
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


PROJECT_ROOT = Path(__file__).resolve().parent.parent
TEMPLATE_PATH = PROJECT_ROOT / "scheduling-template.xlsx"
LOGO_PATH = PROJECT_ROOT / "assets" / "logo.png"


COLOR_PRIMARY = "0E7C66"
COLOR_PRIMARY_DARK = "0A5E4D"
COLOR_PRIMARY_LIGHT = "B7E0D3"
COLOR_PRIMARY_PALE = "E6F4EF"
COLOR_ALT_ROW = "F4FBF9"
COLOR_STATUS_PENDING = "FFF3CD"
COLOR_STATUS_DONE = "D4EDDA"
COLOR_STATUS_FAIL = "F8D7DA"
COLOR_BORDER = "9FBDB3"
COLOR_GRAY_BG = "E5E7EB"
COLOR_GRAY_TEXT = "6B7280"
COLOR_TEXT_DARK = "1F3D38"
COLOR_TEXT_ALERT = "9C2A2A"


THIN_SIDE = Side(style="thin", color=COLOR_BORDER)
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

TITLE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=18)
GUIDE_TITLE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=28)
SUBTITLE_FONT = Font(name="Calibri", italic=True, color=COLOR_TEXT_DARK, size=11)
GUIDE_SUBTITLE_FONT = Font(name="Calibri", italic=True, color=COLOR_TEXT_DARK, size=12)
BANNER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SECTION_FONT = Font(name="Calibri", bold=True, color=COLOR_TEXT_DARK, size=11)
BODY_FONT = Font(name="Calibri", color=COLOR_TEXT_DARK, size=11)
MUTED_FONT = Font(name="Calibri", color=COLOR_GRAY_TEXT, size=11)
STEP_TITLE_FONT = Font(name="Calibri", bold=True, color=COLOR_PRIMARY_DARK, size=12)
FOOTER_FONT = Font(name="Calibri", italic=True, color=COLOR_TEXT_DARK, size=9)
SUMMARY_FONT = Font(name="Calibri", bold=True, color=COLOR_TEXT_DARK, size=11)

TITLE_FILL = PatternFill("solid", fgColor=COLOR_PRIMARY)
SUBTITLE_FILL = PatternFill("solid", fgColor=COLOR_PRIMARY_LIGHT)
BANNER_FILL = PatternFill("solid", fgColor=COLOR_PRIMARY_DARK)
SECONDARY_BANNER_FILL = PatternFill("solid", fgColor=COLOR_PRIMARY)
HEADER_FILL = PatternFill("solid", fgColor=COLOR_PRIMARY)
COMPUTE_HEADER_FILL = PatternFill("solid", fgColor=COLOR_PRIMARY_DARK)
GRAY_FILL = PatternFill("solid", fgColor=COLOR_GRAY_BG)
ALT_FILL = PatternFill("solid", fgColor=COLOR_ALT_ROW)
PALE_FILL = PatternFill("solid", fgColor=COLOR_PRIMARY_PALE)
DONE_FILL = PatternFill("solid", fgColor=COLOR_STATUS_DONE)
FAIL_FILL = PatternFill("solid", fgColor=COLOR_STATUS_FAIL)
PENDING_FILL = PatternFill("solid", fgColor=COLOR_STATUS_PENDING)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)


CLINIC_ROWS = []

MEETING_ROWS = []

DUTY_ROWS = []

DOCTOR_ROWS = []

PATIENT_ROWS = []

PROPOSED_DATES = [
    datetime(2026, 5, 11),
    datetime(2026, 5, 12),
    datetime(2026, 5, 13),
    datetime(2026, 5, 14),
]


def _set_column_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _style_cell(cell, *, font=None, fill=None, alignment=None, border=None, number_format=None):
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    if number_format is not None:
        cell.number_format = number_format


def _merged_banner(ws, cell_range, text, *, fill, font, height=24, alignment=CENTER):
    ws.merge_cells(cell_range)
    first = cell_range.split(":")[0]
    cell = ws[first]
    cell.value = text
    cell.fill = fill
    cell.font = font
    cell.alignment = alignment
    row = int("".join(ch for ch in first if ch.isdigit()))
    ws.row_dimensions[row].height = height


def _write_header_cell(ws, coord, value, *, fill=HEADER_FILL, font=HEADER_FONT):
    cell = ws[coord]
    cell.value = value
    cell.font = font
    cell.fill = fill
    cell.alignment = CENTER
    cell.border = THIN_BORDER


def _build_guide(ws):
    _set_column_widths(ws, [4, 22, 22, 22, 22, 22, 14, 14])
    ws.sheet_view.showGridLines = False

    _merged_banner(
        ws,
        "A1:F1",
        "MEDMATE — XẾP LỊCH PHẪU THUẬT TỰ ĐỘNG",
        fill=TITLE_FILL,
        font=GUIDE_TITLE_FONT,
        height=44,
    )
    _merged_banner(
        ws,
        "A2:F2",
        "Hướng dẫn sử dụng template xếp lịch phẫu thuật (4 sheet)",
        fill=SUBTITLE_FILL,
        font=GUIDE_SUBTITLE_FONT,
        height=22,
    )

    ws.row_dimensions[3].height = 8

    steps = [
        (
            "Bước 1: Điền lịch & danh sách bác sĩ",
            "Mở sheet `INPUT`. Các mục 1–5 nằm cạnh nhau theo hàng ngang: "
            "1=Lịch phòng khám, 2=Lịch họp, 3=Lịch trực cấp cứu, 4=Danh sách bác sĩ, 5=Danh sách bệnh nhân.",
        ),
        (
            "Bước 2: Kiểm tra sheet TÍNH TOÁN",
            "Các công thức tự động kiểm tra xung đột lịch, đề xuất bác sĩ + ngày mổ. "
            "Không sửa tay trong sheet này.",
        ),
        (
            "Bước 3: Yêu cầu Codex chạy xếp lịch",
            "Trong Codex Desktop, gõ: \"Chạy xếp lịch tuần này\". Codex sẽ "
            "mở workbook, đọc lại OUTPUT và báo cáo kết quả.",
        ),
        (
            "Bước 4: Xem sheet OUTPUT",
            "Sheet `OUTPUT` liệt kê các ca \"Đủ điều kiện\" đã được xếp; "
            "các ca \"Loại\" cần điều chỉnh INPUT.",
        ),
    ]

    row = 4
    for idx, (title, body) in enumerate(steps):
        fill = PALE_FILL if idx % 2 == 0 else ALT_FILL
        ws.merge_cells(f"B{row}:F{row}")
        title_cell = ws.cell(row=row, column=2, value=title)
        _style_cell(title_cell, font=STEP_TITLE_FONT, fill=fill, alignment=LEFT_WRAP)
        ws.row_dimensions[row].height = 22

        ws.merge_cells(f"B{row + 1}:F{row + 2}")
        body_cell = ws.cell(row=row + 1, column=2, value=body)
        _style_cell(body_cell, font=BODY_FONT, fill=fill, alignment=LEFT_TOP)
        ws.row_dimensions[row + 1].height = 22
        ws.row_dimensions[row + 2].height = 22

        for r in (row, row + 1, row + 2):
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill
        row += 3

    ws.row_dimensions[18].height = 10

    legend_header = ws.cell(row=19, column=2, value="Chú thích trạng thái")
    legend_header.font = STEP_TITLE_FONT
    legend_header.alignment = LEFT_WRAP
    ws.merge_cells("B19:F19")

    legend_rows = [
        ("Đủ điều kiện", COLOR_STATUS_DONE, "Ca mổ đã được xếp đủ điều kiện và sẽ xuất sang OUTPUT."),
        ("Loại", COLOR_STATUS_FAIL, "Ca mổ bị loại do trùng phòng khám, họp, trực hoặc sai loại PT."),
        ("Ca/ngày tối đa", COLOR_STATUS_PENDING, "Mức trần 11 ca mỗi ngày, áp dụng tự động trong công thức."),
    ]
    for offset, (label, color, desc) in enumerate(legend_rows):
        r = 20 + offset
        swatch = ws.cell(row=r, column=2, value=label)
        swatch.fill = PatternFill("solid", fgColor=color)
        swatch.font = SECTION_FONT
        swatch.alignment = CENTER
        swatch.border = THIN_BORDER
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        desc_cell = ws.cell(row=r, column=3, value=desc)
        desc_cell.font = BODY_FONT
        desc_cell.alignment = LEFT_WRAP
        ws.row_dimensions[r].height = 20

    ws.row_dimensions[24].height = 8
    ws.merge_cells("A25:F25")
    footer = ws["A25"]
    footer.value = "Phiên bản template: 2026.05.16 — Chỉ dùng cho mục đích xếp lịch phẫu thuật trong nội bộ."
    footer.font = FOOTER_FONT
    footer.alignment = CENTER

    try:
        from openpyxl.drawing.image import Image as XLImage

        if LOGO_PATH.exists():
            img = XLImage(str(LOGO_PATH))
            img.width = 120
            img.height = 40
            ws.add_image(img, "H1")
    except Exception:
        pass


def _build_input(ws):
    widths = [14, 10, 12, 2, 14, 10, 20, 2, 14, 10, 12, 2, 12, 20, 16, 2, 12, 22, 14]
    _set_column_widths(ws, widths)
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = "1:2"

    _merged_banner(
        ws,
        "A1:S1",
        "MEDMATE — DỮ LIỆU ĐẦU VÀO",
        fill=TITLE_FILL,
        font=TITLE_FONT,
        height=32,
    )
    _merged_banner(
        ws,
        "A2:S2",
        "Nhập dữ liệu lịch phòng khám, họp, trực, bác sĩ và bệnh nhân",
        fill=SUBTITLE_FILL,
        font=SUBTITLE_FONT,
        height=22,
    )

    _merged_banner(ws, "A3:C3", "1 — Lịch phòng khám", fill=BANNER_FILL, font=BANNER_FONT, height=22)
    _merged_banner(ws, "E3:G3", "2 — Lịch họp", fill=BANNER_FILL, font=BANNER_FONT, height=22)
    _merged_banner(ws, "I3:K3", "3 — Lịch trực cấp cứu", fill=BANNER_FILL, font=BANNER_FONT, height=22)
    _merged_banner(ws, "M3:O3", "4 — Danh sách bác sĩ", fill=BANNER_FILL, font=BANNER_FONT, height=22)
    _merged_banner(
        ws, "Q3:S3", "5 — Danh sách bệnh nhân / ca cần xếp",
        fill=BANNER_FILL, font=BANNER_FONT, height=22,
    )

    for coord, value in (
        ("A4", "Ngày"), ("B4", "Thứ"), ("C4", "Mã BS"),
        ("E4", "Ngày"), ("F4", "Thứ"), ("G4", "Nội dung"),
        ("I4", "Ngày"), ("J4", "Thứ"), ("K4", "Mã BS"),
        ("M4", "Mã BS"), ("N4", "Tên bác sĩ"), ("O4", "Loại PT phụ trách"),
        ("Q4", "Mã BN"), ("R4", "Tên bệnh nhân"), ("S4", "Loại PT"),
    ):
        _write_header_cell(ws, coord, value)
    ws.row_dimensions[4].height = 22
    ws.freeze_panes = "A5"

    for offset in range(36):
        r = 5 + offset

        if offset < len(CLINIC_ROWS):
            date, ma_bs = CLINIC_ROWS[offset]
            ws.cell(r, 1, date).number_format = "dd/mm/yyyy"
            ws.cell(r, 2, f'=TEXT(A{r},"[$-vi-VN]ddd")')
            ws.cell(r, 3, ma_bs)

        if offset < len(MEETING_ROWS):
            date, content = MEETING_ROWS[offset]
            ws.cell(r, 5, date).number_format = "dd/mm/yyyy"
            ws.cell(r, 6, f'=TEXT(E{r},"[$-vi-VN]ddd")')
            ws.cell(r, 7, content)

        if offset < len(DUTY_ROWS):
            date, ma_bs = DUTY_ROWS[offset]
            ws.cell(r, 9, date).number_format = "dd/mm/yyyy"
            ws.cell(r, 10, f'=TEXT(I{r},"[$-vi-VN]ddd")')
            ws.cell(r, 11, ma_bs)

        if offset < len(DOCTOR_ROWS):
            code, name, pt = DOCTOR_ROWS[offset]
            ws.cell(r, 13, code)
            ws.cell(r, 14, name)
            ws.cell(r, 15, pt)

        if offset < len(PATIENT_ROWS):
            ma_bn, ten, loai_pt = PATIENT_ROWS[offset]
            ws.cell(r, 17, ma_bn)
            ws.cell(r, 18, ten)
            ws.cell(r, 19, loai_pt)

        for col in (1, 2, 3):
            cell = ws.cell(r, col)
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            if (r % 2) == 1:
                cell.fill = ALT_FILL

        for col in (5, 6, 7):
            cell = ws.cell(r, col)
            cell.font = BODY_FONT
            cell.alignment = CENTER if col != 7 else LEFT_WRAP
            cell.border = THIN_BORDER
            if (r % 2) == 1:
                cell.fill = ALT_FILL

        for col in (9, 10, 11):
            cell = ws.cell(r, col)
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            if (r % 2) == 1:
                cell.fill = ALT_FILL

        for col in (13, 14, 15):
            cell = ws.cell(r, col)
            cell.font = BODY_FONT
            cell.alignment = CENTER if col != 14 else LEFT_WRAP
            cell.border = THIN_BORDER
            if (r % 2) == 1:
                cell.fill = ALT_FILL

        for col in (17, 18, 19):
            cell = ws.cell(r, col)
            cell.font = BODY_FONT
            cell.alignment = CENTER if col != 18 else LEFT_WRAP
            cell.border = THIN_BORDER
            if (r % 2) == 1:
                cell.fill = ALT_FILL

        ws.row_dimensions[r].height = 18

    pt_dv = DataValidation(
        type="list",
        formula1='"Ghép gan,Cắt gan,Tái thùy,Cắt gan S"',
        allow_blank=True,
    )
    pt_dv.error = "Chọn Loại PT từ danh sách"
    pt_dv.errorTitle = "Loại PT không hợp lệ"
    pt_dv.prompt = "Ghép gan / Cắt gan / Tái thùy / Cắt gan S"
    pt_dv.promptTitle = "Loại PT"
    ws.add_data_validation(pt_dv)
    pt_dv.add("S5:S40")

    ws.auto_filter.ref = "Q4:S40"


def _build_tinh_toan(ws):
    widths = [10, 16, 12, 16, 20, 14, 12, 18, 12, 12, 16, 10, 18, 14]
    _set_column_widths(ws, widths)
    ws.sheet_view.showGridLines = False

    _merged_banner(
        ws,
        "A1:N1",
        "KHU VỰC TÍNH TOÁN (KHÔNG SỬA TAY)",
        fill=GRAY_FILL,
        font=Font(name="Calibri", bold=True, color=COLOR_TEXT_DARK, size=12),
        height=22,
    )
    _merged_banner(
        ws,
        "A2:N2",
        "Các công thức tự động tính toán dựa trên dữ liệu INPUT. Không sửa tay.",
        fill=GRAY_FILL,
        font=Font(name="Calibri", italic=True, color=COLOR_TEXT_DARK, size=11),
        height=20,
    )
    ws.row_dimensions[3].height = 8

    ws["A1"].comment = Comment(
        "Khu vực tự động: công thức tham chiếu INPUT! Không nhập tay.",
        "MedMate",
    )

    compute_headers = [
        "Mã BN", "Tên BN", "Loại PT", "Mã BS dự kiến", "Tên BS",
        "Ngày dự kiến", "Thứ", "Check phòng khám", "Check họp",
        "Check trực", "Check loại PT", "Tổng OK", "Số ca trong ngày", "Kết luận",
    ]
    for col_idx, header in enumerate(compute_headers, start=1):
        cell = ws.cell(row=64, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = COMPUTE_HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[64].height = 28

    for idx in range(36):
        r = 65 + idx
        patient_row = 5 + idx
        ws.cell(r, 1, f"=INPUT!Q{patient_row}")
        ws.cell(r, 2, f"=INPUT!R{patient_row}")
        ws.cell(r, 3, f"=INPUT!S{patient_row}")
        ws.cell(
            r,
            4,
            (
                f'=IF(C{r}="Ghép gan",'
                f'IF(MOD(COUNTIF($C$65:C{r},"Ghép gan"),2)=1,"BS04","BS02"),'
                f'IF(C{r}="Cắt gan",'
                f'IF(MOD(COUNTIF($C$65:C{r},"Cắt gan"),2)=1,"BS05","BS09"),'
                f'IF(C{r}="Tái thùy","BS08",'
                f'IF(C{r}="Cắt gan S","BS06",""))))'
            ),
        )
        ws.cell(r, 5, f'=XLOOKUP(D{r},INPUT!$M$5:$M$10,INPUT!$N$5:$N$10,"")')
        date_cell = ws.cell(r, 6, PROPOSED_DATES[idx % 4])
        date_cell.number_format = "dd/mm/yyyy"
        ws.cell(r, 7, f'=TEXT(F{r},"[$-vi-VN]dddd")')
        ws.cell(r, 8, f'=IF(COUNTIFS(INPUT!$A$5:$A$11,F{r},INPUT!$C$5:$C$11,D{r})>0,"Bận phòng khám","OK")')
        ws.cell(r, 9, f'=IF(COUNTIF(INPUT!$E$5:$E$11,F{r})>0,"Bận họp","OK")')
        ws.cell(r, 10, f'=IF(COUNTIFS(INPUT!$I$5:$I$10,F{r},INPUT!$K$5:$K$10,D{r})>0,"Bận trực","OK")')
        ws.cell(r, 11, f'=IF(XLOOKUP(D{r},INPUT!$M$5:$M$10,INPUT!$O$5:$O$10,"")=C{r},"OK","Sai loại PT")')
        ws.cell(r, 12, f'=IF(AND(H{r}="OK",I{r}="OK",J{r}="OK",K{r}="OK"),"OK","Bận lịch")')
        ws.cell(r, 13, f"=COUNTIF($F$65:$F$100,F{r})")
        ws.cell(r, 14, f'=IF(AND(L{r}="OK",M{r}<=11),"Đủ điều kiện","Loại")')
        for col in range(1, 15):
            cell = ws.cell(r, col)
            cell.font = MUTED_FONT
            cell.alignment = CENTER if col not in (2, 5) else LEFT_WRAP
            cell.border = THIN_BORDER

    verdict_range = "N65:N100"
    ws.conditional_formatting.add(
        verdict_range,
        CellIsRule(
            operator="equal",
            formula=['"Đủ điều kiện"'],
            fill=DONE_FILL,
            font=Font(color=COLOR_TEXT_DARK, bold=True),
        ),
    )
    ws.conditional_formatting.add(
        verdict_range,
        CellIsRule(
            operator="equal",
            formula=['"Loại"'],
            fill=FAIL_FILL,
            font=Font(color=COLOR_TEXT_ALERT, bold=True),
        ),
    )


def _build_output(ws):
    widths = [14, 12, 24, 28, 18, 16]
    _set_column_widths(ws, widths)
    ws.sheet_view.showGridLines = False

    _merged_banner(
        ws,
        "A1:F1",
        "OUTPUT CUỐI CÙNG — DANH SÁCH NGÀY MỔ ĐÃ XẾP",
        fill=SECONDARY_BANNER_FILL,
        font=Font(name="Calibri", bold=True, color="FFFFFF", size=14),
        height=26,
    )

    output_headers = ["Ngày mổ", "Thứ", "Bác sĩ", "Bệnh nhân", "Loại PT", "Kết luận"]
    for col_idx, header in enumerate(output_headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = COMPUTE_HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[2].height = 24

    for idx in range(36):
        r = 3 + idx
        src = 65 + idx
        ws.cell(r, 1, f'=IF(TÍNH TOÁN!$N{src}="Đủ điều kiện",TEXT(TÍNH TOÁN!$F{src},"yyyy-mm-dd"),"")')
        ws.cell(r, 2, f'=IF(TÍNH TOÁN!$N{src}="Đủ điều kiện",TÍNH TOÁN!$G{src},"")')
        ws.cell(r, 3, f'=IF(TÍNH TOÁN!$N{src}="Đủ điều kiện",TÍNH TOÁN!$E{src},"")')
        ws.cell(r, 4, f'=IF(TÍNH TOÁN!$N{src}="Đủ điều kiện",TÍNH TOÁN!$B{src},"")')
        ws.cell(r, 5, f'=IF(TÍNH TOÁN!$N{src}="Đủ điều kiện",TÍNH TOÁN!$C{src},"")')
        ws.cell(r, 6, f"=TÍNH TOÁN!$N{src}")
        for col in range(1, 7):
            cell = ws.cell(r, col)
            cell.font = BODY_FONT
            cell.alignment = CENTER if col not in (3, 4) else LEFT_WRAP
            cell.border = THIN_BORDER
            if (r % 2) == 1:
                cell.fill = ALT_FILL

    output_verdict_range = "F3:F38"
    ws.conditional_formatting.add(
        output_verdict_range,
        CellIsRule(
            operator="equal",
            formula=['"Đủ điều kiện"'],
            fill=DONE_FILL,
            font=Font(color=COLOR_TEXT_DARK, bold=True),
        ),
    )
    ws.conditional_formatting.add(
        output_verdict_range,
        CellIsRule(
            operator="equal",
            formula=['"Loại"'],
            fill=FAIL_FILL,
            font=Font(color=COLOR_TEXT_ALERT, bold=True),
        ),
    )

    ws.auto_filter.ref = "A2:F2"
    ws.freeze_panes = "A3"

    summary_pairs = [
        ("A40", "Tổng ca đủ điều kiện", DONE_FILL),
        ("B40", '=COUNTIF(TÍNH TOÁN!$N$65:$N$100,"Đủ điều kiện")', DONE_FILL),
        ("C40", "Tổng ca bị loại", FAIL_FILL),
        ("D40", '=COUNTIF(TÍNH TOÁN!$N$65:$N$100,"Loại")', FAIL_FILL),
        ("F40", "Ca/ngày tối đa", PENDING_FILL),
        ("G40", 11, PENDING_FILL),
    ]
    for coord, value, fill in summary_pairs:
        cell = ws[coord]
        cell.value = value
        cell.font = SUMMARY_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[40].height = 24


def build() -> Path:
    wb = Workbook()
    ws_guide = wb.active
    ws_guide.title = "HƯỚNG DẪN"
    ws_input = wb.create_sheet("INPUT")
    ws_compute = wb.create_sheet("TÍNH TOÁN")
    ws_output = wb.create_sheet("OUTPUT")

    _build_guide(ws_guide)
    _build_input(ws_input)
    _build_tinh_toan(ws_compute)
    _build_output(ws_output)

    if TEMPLATE_PATH.exists():
        os.remove(TEMPLATE_PATH)
    wb.save(TEMPLATE_PATH)
    return TEMPLATE_PATH


def main() -> int:
    path = build()
    print(f"Built {path.name} with sheets: HƯỚNG DẪN, INPUT, TÍNH TOÁN, OUTPUT")
    return 0


if __name__ == "__main__":
    sys.exit(main())
